
import csv
import sys
from openpyxl import Workbook, load_workbook
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtWidgets import *
from mailmerge import MailMerge

class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Appmerge")
        Form.resize(1200, 690)
        icn = QtGui.QIcon('icon.ico')
        Form.setWindowIcon(icn)
        font = QtGui.QFont()
        font.setFamily("Open Sans")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        lb = QtWidgets.QLabel(Form)
        lb.setStyleSheet("border: 2px inset grey; color: grey")
        lb.setText('Deve By CHAHBOUN Mohammed')
        lb.setGeometry(QtCore.QRect(1000, 640, 360, 50))
        lb.setFont(font)
        self.c = QtWidgets.QComboBox(Form)
        self.c.setGeometry(QtCore.QRect(750, 20, 261, 31))
        self.c.setObjectName("c")
        self.rech = QtWidgets.QPushButton(Form)
        self.rech.setGeometry(QtCore.QRect(440, 20, 121, 31))
        self.rech.setObjectName("rech")
        self.tbl = QtWidgets.QTableWidget(Form)
        self.tbl.setGeometry(QtCore.QRect(5, 71, 1355, 551))
        self.tbl.setObjectName("tbl")
        self.enr = QtWidgets.QPushButton(Form)
        self.enr.setGeometry(QtCore.QRect(1030, 20, 121, 31))
        self.enr.setObjectName("enr")
        self.ml = QtWidgets.QPushButton(Form)
        self.ml.setGeometry(QtCore.QRect(1200, 20, 121, 31))
        self.ml.setObjectName("ml")
        self.m = QtWidgets.QLineEdit(Form)
        self.m.setGeometry(QtCore.QRect(150, 20, 271, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.m.setFont(font)
        self.m.setObjectName("m")
        self.sl = QtWidgets.QPushButton(Form)
        self.sl.setGeometry(QtCore.QRect(10, 20, 121, 31))
        self.sl.setObjectName("sl")
        self.lb = QtWidgets.QLabel(Form)
        self.lb.setGeometry(QtCore.QRect(580, 19, 161, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(False)
        font.setWeight(50)
        self.lb.setFont(font)
        self.lb.setObjectName("lb")
        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "appmerge"))
        self.rech.setText(_translate("Form", "Recherche"))
        self.enr.setText(_translate("Form", "Enregistrer"))
        self.ml.setText(_translate("Form", "Mail Merge"))
        self.sl.setText(_translate("Form", "Sélectionner xlsx"))
        self.lb.setText(_translate("Form", "Critière de Recherche :"))


class App_Window(QMainWindow, Ui_Form):
    def __init__(self):
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.st()

    def st(self):
        QApplication.setStyle(QStyleFactory.create('Fusion'))
        self.sl.clicked.connect(self.ajt)
        self.rech.clicked.connect(self.recherche)
        self.enr.clicked.connect(self.enregistrer)
        self.ml.clicked.connect(self.mail)
        self.rech.setEnabled(False)
        self.enr.setEnabled(False)
        self.ml.setEnabled(False)
        self.c.setEnabled(False)

    def ajt(self):
        file = QFileDialog.getOpenFileName(self, 'Open file', 'c:\\', ("Excel file (*.xlsx);;CSV file (*.csv)"))
        self.f = file[0]
        #a = requests.request(method='get', url='https://www.facebook.com/permalink.php?story_fbid=212770222835731&id=100023082630505')
        #if 'MNI SERVICE' not in a.text:
        self.rech.setEnabled(True)
        self.c.setEnabled(True)
        self.enr.setEnabled(True)
        self.ml.setEnabled(True)
        '''else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText("Error")
            msg.setInformativeText("Contact MNI SERVICES")
            msg.setWindowTitle("ERROR")
            msg.setDetailedText("MORE INFO VISITE www.mniservices.ma")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
            self.rech.setEnabled(False)
            self.c.setEnabled(False)
            self.enr.setEnabled(False)'''
        self.l = 'abcdefghijklmnopqrstuvwxyz'
        self.lis = []
        if 'xlsx' in self.f:
            book = load_workbook(self.f)
            sh = book.sheetnames
            a = book[sh[0]]
            j = 0
            for i in a['1']:
                q = a['{}1'.format(self.l[j])].value
                self.lis.append(q)
                j += 1
            self.c.addItems(self.lis)
            self.tbl.setColumnCount(len(self.lis))
        elif 'csv' in self.f:
            with open(self.f, "r") as f:
                red = csv.DictReader(f)
                headers = red.fieldnames
                for i in headers:
                    self.lis = str(i).split(';')
            self.c.addItems(self.lis)
            self.tbl.setColumnCount(len(self.lis))

    def recherche(self):
        if 'xlsx' in self.f:
            book = load_workbook(self.f)
            sh = book.sheetnames
            a = book[sh[0]]
            cr = self.c.currentIndex()
            self.tbl.setHorizontalHeaderLabels(self.lis)
            q = 0
            r = 0
            n = 1
            for i in a['a']:
                b = a['{}{}'.format(str(self.l[cr]), str(n))]
                for w in self.lis:
                    if str(b.value).lower()[0:len(str(self.m.text()))] == str(self.m.text()).lower():
                        c = a['{}{}'.format(str(self.l[q]), str(n))]
                        self.tbl.setRowCount(1 + r)
                        if str(c.value) == 'None':
                            self.tbl.setItem(r, q, QTableWidgetItem(''))
                        else:
                            self.tbl.setItem(r, q, QTableWidgetItem(str(c.value)))
                        q += 1
                        if q >= len(self.lis):
                            r += 1
                q = 0
                n += 1
        elif 'csv' in self.f:
            t = 0
            x = 0
            j = self.c.currentText()
            self.tbl.setHorizontalHeaderLabels(self.lis)
            with open(self.f, mode='r') as f:
                reader = csv.DictReader(f, delimiter=';')
                w = 0
                for i in reader:
                    if str(i[j]).lower()[0:len(str(self.m.text()))] == str(self.m.text()).lower():
                        self.tbl.setRowCount(1 + w)
                        w += 1
                        for o in self.lis:
                            self.tbl.setItem(t, x, QTableWidgetItem(str(i[o])))
                            x += 1
                t += 1

    def enregistrer(self):
        file = QFileDialog.getSaveFileName(self, 'Enregistrer Fichier', str(self.m.text()),
                                           ("Excel file (*.xlsx);;CSV file (*.csv)"))
        wb = Workbook()
        ws = wb.active
        for u in range(len(self.lis)):
            ws['{}1'.format(str(self.l[u]))] = self.lis[u]
        for w in range(self.tbl.rowCount()):
            for i in range(len(self.lis)):
                try:
                    ws['{}{}'.format(str(self.l[i]), str(w + 2))] = self.tbl.item(w, i).text()
                except Exception:
                    pass
        wb.save(file[0])

    def mail(self):
        template = "res/template.docx"
        document = MailMerge(template)
        file = QFileDialog.getSaveFileName(self, 'Enregistrer Fichier', str(self.m.text()),
                                           ("Word file (*.docx)"))
        w = document.get_merge_fields()
        w = sorted(w)
        d = {}
        h = [0,1,2,3,4,5]
        li = []
        for x in range(self.tbl.rowCount()):
            for i in range(len(h)):
                d[w[i]] = self.tbl.item(x, h[i]).text()
            li.append(dict(d))
        document.merge_pages(li)
        document.write(file[0])


if __name__ == '__main__':
    app = QApplication(sys.argv)
    frm = App_Window()
    frm.showMaximized()
    sys.exit(app.exec_())
